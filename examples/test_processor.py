#!/usr/bin/env python3
"""
India Property Hunt — Example Test Processor (Bangalore)
Demonstrates parsing listings, applying India-specific priority logic,
updating tracker spreadsheet, generating outreach files, and building email summary.

Usage:
  pip install openpyxl
  python3 examples/test_processor.py
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime, date
import os
import re

# ─── CONFIG ───────────────────────────────────────────────────────────────
HUNT_DIR = "."
TRACKER_FILE = os.path.join(HUNT_DIR, "india_property_hunt.xlsx")
OUTREACH_DIR = os.path.join(HUNT_DIR, "outreach")
TODAY = date.today().strftime("%Y-%m-%d")
NOW = datetime.now().strftime("%H:%M")

PRIMARY_AREAS = ["hsr layout", "bellandur", "koramangala", "indiranagar"]
SECONDARY_AREAS = ["sarjapur road", "marathahalli", "whitefield", "electronic city"]
TERTIARY_AREAS = ["hebbal", "jp nagar", "jayanagar", "yelahanka"]

BUDGET_1BHK = 28000
BUDGET_2BHK = 45000
BUDGET_3BHK = 65000
MAX_DEPOSIT_MONTHS = 6
MAX_MAINTENANCE = 5000
MOVE_IN_DATE = "2026-06-01"

PRIORITY_HIGH_FILL = PatternFill("solid", fgColor="E2EFDA")
PRIORITY_MED_FILL = PatternFill("solid", fgColor="FFFFC7")
PRIORITY_LOW_FILL = PatternFill("solid", fgColor="FCE4D6")

# ─── EXAMPLE LISTINGS ────────────────────────────────────────────────────
# These are representative listings found via NoBroker, Housing.com, etc.

listings = [
    {
        "title": "2BHK Semi-furnished, Silver Nest, Haralur",
        "platform": "NoBroker",
        "url": "https://www.nobroker.in/property/2-bhk-apartment-for-rent-in-hsr-layout-bangalore",
        "area": "HSR Layout",
        "bhk": "2BHK",
        "rent": 40000,
        "deposit": 240000,
        "deposit_months": 6,
        "maintenance": 3500,
        "brokerage": 0,
        "available": "Immediate",
        "furnishing": "Semi-furnished",
        "area_sqft": 1100,
        "water": "Cauvery",
        "power_backup": "Inverter",
        "parking": "Bike",
        "amenities": "Lift, Security, CCTV, Garden",
        "society": "Silver Nest",
        "construction_age": "New",
        "bachelor_friendly": "Yes",
        "notes": "Near HSR BDA Complex, walking distance to restaurants.",
    },
    {
        "title": "1BHK Fully furnished, Marthahalli",
        "platform": "NoBroker",
        "url": "https://nobroker.in/property/1-bhk-apartment-for-rent-in-marthahalli",
        "area": "Marathahalli",
        "bhk": "1BHK",
        "rent": 26000,
        "deposit": 156000,
        "deposit_months": 6,
        "maintenance": 2000,
        "brokerage": 0,
        "available": "Immediate",
        "furnishing": "Fully furnished",
        "area_sqft": 650,
        "water": "Borewell",
        "power_backup": "Inverter",
        "parking": "Bike",
        "amenities": "Lift, Security",
        "society": "Apartment Complex",
        "construction_age": "<5 years",
        "bachelor_friendly": "Yes",
        "notes": "Near RMZ Ecospace and ORR.",
    },
    {
        "title": "2BHK Semi-furnished, Elv Marvel, Whitefield",
        "platform": "NoBroker",
        "url": "https://nobroker.in/property/2-bhk-apartment-for-rent-in-whitefield",
        "area": "Whitefield",
        "bhk": "2BHK",
        "rent": 43000,
        "deposit": 258000,
        "deposit_months": 6,
        "maintenance": 4500,
        "brokerage": 0,
        "available": "1 May 2026",
        "furnishing": "Semi-furnished",
        "area_sqft": 1200,
        "water": "Cauvery",
        "power_backup": "Genset",
        "parking": "Car + Bike",
        "amenities": "Gym, Swimming Pool, Clubhouse, Security",
        "society": "Elv Marvel",
        "construction_age": "New",
        "bachelor_friendly": "Yes",
        "notes": "Premium gated community near ITPL.",
    },
    {
        "title": "2BHK Semi-furnished, Koramangala 5th Block",
        "platform": "Housing.com",
        "url": "https://housing.com/rent/2bhk-koramangala",
        "area": "Koramangala",
        "bhk": "2BHK",
        "rent": 55000,
        "deposit": 330000,
        "deposit_months": 6,
        "maintenance": 5000,
        "brokerage": 55000,
        "available": "15 May 2026",
        "furnishing": "Semi-furnished",
        "area_sqft": 950,
        "water": "Cauvery",
        "power_backup": "Inverter",
        "parking": "Bike",
        "amenities": "Lift, Security, CCTV",
        "society": "Independent House",
        "construction_age": "10+ years",
        "bachelor_friendly": "Yes",
        "notes": "Prime Koramangala. Has brokerage. Old building.",
    },
    {
        "title": "2BHK, Bellandur (via broker)",
        "platform": "Housing.com",
        "url": "https://housing.com/rent/2bhk-bellandur",
        "area": "Bellandur",
        "bhk": "2BHK",
        "rent": 45000,
        "deposit": 270000,
        "deposit_months": 6,
        "maintenance": 4500,
        "brokerage": 45000,
        "available": "Immediate",
        "furnishing": "Semi-furnished",
        "area_sqft": 1500,
        "water": "Cauvery",
        "power_backup": "Genset",
        "parking": "Car",
        "amenities": "Gym, Pool, Clubhouse, Garden",
        "society": "Gated Community",
        "construction_age": "<5 years",
        "bachelor_friendly": "Yes",
        "notes": "Spacious 2BHK. ₹45K brokerage.",
    },
    {
        "title": "1BHK Fully furnished, Damden Aalada Mara, Electronic City",
        "platform": "NoBroker",
        "url": "https://nobroker.in/property/1-bhk-electronic-city",
        "area": "Electronic City",
        "bhk": "1BHK",
        "rent": 29000,
        "deposit": 174000,
        "deposit_months": 6,
        "maintenance": 3000,
        "brokerage": 0,
        "available": "Immediate",
        "furnishing": "Fully furnished",
        "area_sqft": 600,
        "water": "Cauvery",
        "power_backup": "Genset",
        "parking": "Car + Bike",
        "amenities": "Gym, Pool, Garden, Clubhouse",
        "society": "Damden Aalada Mara",
        "construction_age": "New",
        "bachelor_friendly": "Yes",
        "notes": "Excellent amenities for the price.",
    },
]

# ─── PRIORITY LOGIC ──────────────────────────────────────────────────────

def get_area_tier(area_lower):
    for a in PRIMARY_AREAS:
        if a in area_lower:
            return "primary"
    for a in SECONDARY_AREAS:
        if a in area_lower:
            return "secondary"
    for a in TERTIARY_AREAS:
        if a in area_lower:
            return "tertiary"
    return "outer"

def get_budget(l):
    bhk = l["bhk"]
    if "1BHK" in bhk:
        return BUDGET_1BHK
    elif "2BHK" in bhk:
        return BUDGET_2BHK
    elif "3BHK" in bhk:
        return BUDGET_3BHK
    return BUDGET_2BHK

def prioritize(l):
    reasons = []
    area_tier = get_area_tier(l["area"].lower())
    budget = get_budget(l)
    
    if l.get("power_backup", "").lower() in ("none", ""):
        return "SKIP", "No power backup"
    if l.get("bachelor_friendly", "").lower() == "no":
        return "SKIP", "Not bachelor friendly"
    
    over_budget = l["rent"] > budget
    if over_budget:
        reasons.append(f"over budget (₹{l['rent']:,} vs ₹{budget:,})")
    if l.get("deposit_months", 0) > MAX_DEPOSIT_MONTHS:
        reasons.append(f"high deposit ({l['deposit_months']} months)")
    if l.get("maintenance", 0) > MAX_MAINTENANCE:
        reasons.append(f"high maintenance (₹{l['maintenance']:,})")
    if l.get("brokerage", 0) > 0:
        reasons.append(f"brokerage ₹{l['brokerage']:,}")
    if l.get("water", "").lower() == "tanker":
        reasons.append("tanker water")
    
    if not over_budget and area_tier == "primary" and l.get("brokerage", 0) == 0:
        priority = "High"
    elif area_tier == "secondary" and not over_budget:
        priority = "Medium"
    elif over_budget:
        priority = "Low"
    else:
        priority = "Medium"
    
    return priority, "; ".join(reasons) if reasons else "good match"

# ─── RUN ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"🏠 India Property Hunt — Example Test Run")
    print(f"📅 {TODAY} | {NOW} UTC")
    print(f"{'='*60}")
    
    results = []
    for l in listings:
        priority, reason = prioritize(l)
        l["priority"] = priority
        l["reason"] = reason
        results.append(l)
        action = "✅" if priority != "SKIP" else "❌"
        print(f"  {action} [{priority:>6}] ₹{l['rent']:>6,} | {l['area']:<20} | {l['bhk']} | {reason}")
    
    high = [l for l in results if l["priority"] == "High"]
    med = [l for l in results if l["priority"] == "Medium"]
    low = [l for l in results if l["priority"] in ("Low", "SKIP")]
    print(f"\n📊 HIGH: {len(high)} | MEDIUM: {len(med)} | LOW/SKIP: {len(low)}")
    print(f"✅ Example run complete!")
    print(f"   In production, this would update tracker, generate outreach, and send email.")